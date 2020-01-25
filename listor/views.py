from django.shortcuts import render

# Create your views here.


import os

# My only view

import threading
from django.http import HttpResponse

def load(request):
    memdata = get_memdata()['data']
#     mk_listor(memdata)
    t = threading.Thread(target=mk_listor, args=(memdata,))
    t.start()
    resptext = "Loaded "+str(len(memdata))+" records!"
    return HttpResponse(resptext)

# Create lists
 
avdelningar = ['Småkrypen', 'Sagodjuren', 'Husdjuren', 'Gosedjuren', 'Fabeldjuren', 'Skogsdjuren', 'Urdjuren', 'Rovdjuren', 'Slow Fox', 'Rover']
grenar = {'Spårare':['Sagodjuren', 'Husdjuren', 'Gosedjuren'], 'Upptäckare':['Fabeldjuren', 'Skogsdjuren'], 'Äventyrare':['Urdjuren', 'Rovdjuren'], 'Utmanare':['Slow Fox']}
# avdelningar = ['Småkrypen','Sagodjuren', 'Husdjuren', 'Fabeldjuren', 'Skogsdjuren', 'Urdjuren', 'Rovdjuren', 'Slow Fox']
# grenar = {'Spårare':['Sagodjuren', 'Husdjuren'], 'Upptäckare':['Fabeldjuren', 'Skogsdjuren'], 'Äventyrare':['Urdjuren', 'Rovdjuren'], 'Utmanare':['Slow Fox']}


import uuid, copy
def mk_listor(memdata):
    kontaktlista(memdata)
    telefonlista(memdata)
    allepost(memdata)

    extra_leaders = {
        "3291427": {"VEM":"Hanna Norström", "VAR":"Urdjuren"},
        "3258930": {"VEM":"Dick Börnesson", "VAR":"Slow Fox"},
        "3239306": {"VEM":"Ida Rosengren", "VAR":"Fabeldjuren"},
        "3254565": {"VEM":"Johannes Moberg", "VAR":"Gosedjuren"},
        "3262806": {"VEM":"Simon Vallin", "VAR":"Gosedjuren"},
        "2001383": {"VEM":"Oskar Grönman", "VAR":"Gosedjuren"},
        "3275813": {"VEM":"Miamaria Broberg", "VAR":"Gosedjuren"},
        "3226730": {"VEM":"Isak Ahlberg", "VAR":"Gosedjuren"},
        "3202688": {"VEM":"Ludvig Jogmark", "VAR":"Rovdjuren"}
        }
    for mx in extra_leaders:
        xl = copy.deepcopy(memdata[mx])
        xl['unit']['value'] = extra_leaders[mx]["VAR"]  # extra unit
        xl['date_of_birth']['value'] = "1990-01-01"    # make a leader
        memdata[uuid.uuid4().hex] = xl      # random member ID

    avdelningslistor(memdata)
    grenlistor(memdata)
    ledarlista(memdata)
    
#     wsj19lista(memdata)
#     sommar19lista(memdata)
# #     testlista(memdata)

def testlista(memdata):
    def v(m,f):
        return memdata[m][f]['value'] if f in memdata[m] else ""

    mlist = [m for m in memdata]
    elista = ""
    for m in mlist:
        elista += v(m,'member_no')+"\n"
    save_file("testlista.txt",elista.encode(encoding="utf-8", errors="strict"))

def avdelningslistor(memdata):
    def v(m,f):
        return memdata[m][f]['value'] if f in memdata[m] else ""

    for avd in avdelningar:
        mlist = [m for m in memdata if memdata[m]['unit']['value'] == avd]
        elista = 'LÄGG ALLTID ALLA E-POSTADRESSER SOM "HEMLIG KOPIA" (bcc) NÄR DU ANVÄNDER E-POSTADRESSERNA I DENNA FIL! GDPR!!!\n'
        for m in mlist:
            namn = v(m,'first_name')+" "+v(m,'last_name')
            if v(m,'email') != "": 
                elista += namn+" <"+v(m,'email')+">;\n"
            if v(m,'contact_email_mum') != "" and v(m,'contact_email_mum') != v(m,'email'):
                elista += v(m,'contact_mothers_name')+" ("+namn+"s anhörig) <"+v(m,'contact_email_mum')+">;\n"
            if v(m,'contact_email_dad') != "" and v(m,'contact_email_dad') != v(m,'email'):
                elista += v(m,'contact_fathers_name')+" ("+namn+"s anhörig) <"+v(m,'contact_email_dad')+">;\n"
            if v(m,'contact_alt_email') != "":
                elista += namn+" (Extra) <"+v(m,'contact_alt_email')+">;\n"
        save_file(avd+".txt",elista.encode(encoding="utf-8", errors="strict"))

def grenlistor(memdata):
    def v(m,f):
        return memdata[m][f]['value'] if f in memdata[m] else ""

    for gren in ['Spårare','Upptäckare','Äventyrare']:
        elista = 'LÄGG ALLTID ALLA E-POSTADRESSER SOM "HEMLIG KOPIA" (bcc) NÄR DU ANVÄNDER E-POSTADRESSERNA I DENNA FIL! GDPR!!!\n'
        for avd in grenar[gren]:
            mlist = [m for m in memdata if memdata[m]['unit']['value'] == avd]
            for m in mlist:
                namn = v(m,'first_name')+" "+v(m,'last_name')
                if v(m,'email') != "": 
                    elista += namn+" <"+v(m,'email')+">;\n"
                if v(m,'contact_email_mum') != "" and v(m,'contact_email_mum') != v(m,'email'):
                    elista += v(m,'contact_mothers_name')+" ("+namn+"s anhörig) <"+v(m,'contact_email_mum')+">;\n"
                if v(m,'contact_email_dad') != "" and v(m,'contact_email_dad') != v(m,'email'):
                    elista += v(m,'contact_fathers_name')+" ("+namn+"s anhörig) <"+v(m,'contact_email_dad')+">;\n"
                if v(m,'contact_alt_email') != "":
                    elista += namn+" (Extra) <"+v(m,'contact_alt_email')+">;\n"
        save_file(gren+".txt",elista.encode(encoding="utf-8", errors="strict"))

def allepost(memdata):
    def v(m,f):
        return memdata[m][f]['value'] if f in memdata[m] else ""

    mlist = [m for m in memdata if memdata[m]['unit']['value'] != "Under avveckling"]
    elist = set()
    for m in mlist:
        if v(m,'email') != "": 
            elist.add(v(m,'email'))
        if v(m,'contact_email_mum') != "" :
            elist.add(v(m,'contact_email_mum'))
        if v(m,'contact_email_dad') != "":
            elist.add(v(m,'contact_email_dad'))
        if v(m,'contact_alt_email') != "":
            elist.add(v(m,'contact_alt_email'))
    data = 'E-POSTADRESSERNA I DENNA FIL FÅR *ENDAST* ANVÄNDAS FÖR UTSKICK SOM ÄR GODKÄNDA AV STYRELSEN\nLÄGG DESSUTOM ALLTID ALLA E-POSTADRESSER SOM "HEMLIG KOPIA" (bcc)! GDPR!!!\n'
    for l in elist:
        data += l+";\n"
    save_file("Alla.txt",data.encode(encoding="utf-8", errors="strict"))

def wsj19lista(memdata):
    def v(m,f):
        return memdata[m][f]['value'] if f in memdata[m] else ""

    deltlist = ["2001383","3209236","3209818","3223010","3224174","3225004","3225577","3227897","3230994","3231772",
                "3232356","3234094","3234250","3235445","3236796","3242159","3249086","3249088","3249444","3254565",
                "3261824","3262806","3273751","3275149","3276352","3280061","3287727","3291282","3291638"]
    
    elist = set()
    for m in deltlist:
        if v(m,'email') != "": 
            elist.add(v(m,'email'))
        if v(m,'contact_email_mum') != "" :
            elist.add(v(m,'contact_email_mum'))
        if v(m,'contact_email_dad') != "":
            elist.add(v(m,'contact_email_dad'))
        if v(m,'contact_alt_email') != "":
            elist.add(v(m,'contact_alt_email'))
    data = ""
    for l in elist:
        data += l+";\n"
    save_file("wsj19.txt",data.encode(encoding="utf-8", errors="strict"))

def sommar19lista(memdata):
    def v(m,f):
        return memdata[m][f]['value'] if f in memdata[m] else ""

    deltlist = ["3237511","3249106","3258930","3275653","3264482","3275606","3323929","3252351","3236153","3291427","3239306","3277549","3295867",
                "3275150","3252064","3291380","3277616","3318013","3325804","3257963","3236899","3234914","3264321","3257965","3236156",
                "3261071","3275828","3234913","3277212","3252118","3274216","3252127","3306182","3280101","3317638","3263786","3264106","3305079","3264317",
                "3333017","3264514","3180862","3197850","3307450","3180762"]
    
    elist = set()
    for m in deltlist:
        if v(m,'email') != "": 
            elist.add(v(m,'email'))
        if v(m,'contact_email_mum') != "" :
            elist.add(v(m,'contact_email_mum'))
        if v(m,'contact_email_dad') != "":
            elist.add(v(m,'contact_email_dad'))
        if v(m,'contact_alt_email') != "":
            elist.add(v(m,'contact_alt_email'))
    data = ""
    for l in elist:
        data += l+";\n"
    save_file("Klarälven19.txt",data.encode(encoding="utf-8", errors="strict"))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.writer.excel import save_virtual_workbook

def kontaktlista(memdata):
    def v(m,f):
        return memdata[m][f]['value'] if f in memdata[m] else ""

    wb = Workbook()
    ws = wb.active
    for avd in avdelningar:
        if avd == 'Småkrypen':
            mlist = [m for m in memdata if memdata[m]['unit']['value'] == avd]
        else:
            mlist = [m for m in memdata if memdata[m]['unit']['value'] == avd and memdata[m]['date_of_birth']['value'] > "1994-01-01"]
        mlist = sorted(mlist,key=lambda m: v(m,'first_name')+" "+v(m,'last_name'))
        ws.title = avd
        header = ["Namn", "Född", "Adress", "Hemtelefon", "Mobiltelefon", "E-post", "Anhörig1 namn", "Anhörig1 mobil", "Anhörig1 e-post", "Anhörig2 namn", "Anhörig2 mobil", "Anhörig2 e-post", "Extra e-post"]   
        colsizes = [30,7,40,14,14,35,30,17,35,30,17,35,35]
        for col in range(len(header)):
            ws.cell(row=1,column=col+1).value = header[col]
            ws.cell(row=1,column=col+1).font = Font(bold=True)
            ws.cell(row=1,column=col+1).fill = PatternFill("solid", fgColor="FFFF00")
            ws.column_dimensions[chr(65+col)].width = colsizes[col]
    
        r = 2
        for m in mlist:
            ws.cell(row=r,column= 1).value = v(m,'first_name')+" "+v(m,'last_name')
            fodd = v(m,'date_of_birth')
#             ws.cell(row=r,column= 2).value = "20"+fodd[2:4] if avd != "Småkrypen" else "19"+fodd[2:4]   # Very ugly fix...
            ws.cell(row=r,column= 2).value = fodd[:4] if fodd > "2000-01-01" else ""
            ws.cell(row=r,column= 3).value = v(m,'address_1')+", "+v(m,'postcode')+" "+v(m,'town')
            ws.cell(row=r,column= 4).value = v(m,'contact_home_phone')
            ws.cell(row=r,column= 5).value = v(m,'contact_mobile_phone')
            ws.cell(row=r,column= 6).value = v(m,'email')
            ws.cell(row=r,column= 7).value = v(m,'contact_mothers_name')
            ws.cell(row=r,column= 8).value = v(m,'contact_mobile_mum')
            ws.cell(row=r,column= 9).value = v(m,'contact_email_mum')
            ws.cell(row=r,column=10).value = v(m,'contact_fathers_name')
            ws.cell(row=r,column=11).value = v(m,'contact_mobile_dad')
            ws.cell(row=r,column=12).value = v(m,'contact_email_dad')
            ws.cell(row=r,column=13).value = v(m,'contact_alt_email')
            r += 1
        ws = wb.create_sheet()

# Ledare
    mlist = [m for m in memdata if memdata[m]['date_of_birth']['value'] < "2000-01-01" and memdata[m]['unit']['value'] not in ["Under avveckling","Småkrypen"]]
    mlist = sorted(mlist,key=lambda m: v(m,'first_name')+" "+v(m,'last_name'))
    ws.title = 'Ledare'
    header = ["Namn", "Avdelning", "Adress", "Hemtelefon", "Mobiltelefon", "E-post", "Extra e-epost"]   
    colsizes = [30,20,40,14,14,35,35]
    for col in range(len(header)):
        ws.cell(row=1,column=col+1).value = header[col]
        ws.cell(row=1,column=col+1).font = Font(bold=True)
        ws.cell(row=1,column=col+1).fill = PatternFill("solid", fgColor="FFFF00")
        ws.column_dimensions[chr(65+col)].width = colsizes[col]
    r = 2
    for m in mlist:
        ws.cell(row=r,column= 1).value = v(m,'first_name')+" "+v(m,'last_name')
        ws.cell(row=r,column= 2).value = v(m,'unit')
        ws.cell(row=r,column= 3).value = v(m,'address_1')+", "+v(m,'postcode')+" "+v(m,'town')
        ws.cell(row=r,column= 4).value = v(m,'contact_home_phone')
        ws.cell(row=r,column= 5).value = v(m,'contact_mobile_phone')
        ws.cell(row=r,column= 6).value = v(m,'email')
        ws.cell(row=r,column= 7).value = v(m,'contact_alt_email')
        r += 1

    save_file("Kontaktlista.xlsx",save_virtual_workbook(wb))


def ledarlista(memdata):
    def v(m,f):
        return memdata[m][f]['value'] if f in memdata[m] else ""

    wb = Workbook()
    ws = wb.active
    for gren in ['Spårare','Upptäckare','Äventyrare','Utmanare']:
        ws.title = gren
        header = ["Namn", "Avdelning", "Mobiltelefon", "E-post"]
        colsizes = [30,15,14,35]
        for col in range(len(header)):
            ws.cell(row=1,column=col+1).value = header[col]
            ws.cell(row=1,column=col+1).font = Font(bold=True)
            ws.cell(row=1,column=col+1).fill = PatternFill("solid", fgColor="FFFF00")
            ws.column_dimensions[chr(65+col)].width = colsizes[col]
        r = 2
        for avd in grenar[gren]:
            mlist = [m for m in memdata if memdata[m]['unit']['value'] == avd and memdata[m]['date_of_birth']['value'] < "2000-01-01"]
            mlist = sorted(mlist,key=lambda m: v(m,'first_name')+" "+v(m,'last_name'))
            for m in mlist:
                ws.cell(row=r,column= 1).value = v(m,'first_name')+" "+v(m,'last_name')
                ws.cell(row=r,column= 2).value = v(m,'unit')
                ws.cell(row=r,column= 3).value = v(m,'contact_mobile_phone')
                ws.cell(row=r,column= 4).value = v(m,'email')
                r += 1
        ws = wb.create_sheet()
#     wb.remove_sheet(ws)     # Remove empty sheet
    wb.remove(ws)     # Remove empty sheet
    save_file("Avdelningsledarlista.xlsx",save_virtual_workbook(wb))

def telefonlista(memdata):
    def v(m,f):
        return memdata[m][f]['value'] if f in memdata[m] else ""

    mlist = [m for m in memdata]
    mlist = sorted(mlist,key=lambda m: v(m,'first_name')+" "+v(m,'last_name'))
    wb = Workbook()
    ws = wb.active
    ws.title = "Telefonlista"
    header = ["Namn", "Avdelning", "Hemtelefon", "Mobiltelefon", "Anhörig1 mobil", "Anhörig2 mobil"]
    colsizes = [30,20,14,14,17,17]
    for col in range(len(header)):
        ws.cell(row=1,column=col+1).value = header[col]
        ws.cell(row=1,column=col+1).font = Font(bold=True)
        ws.cell(row=1,column=col+1).fill = PatternFill("solid", fgColor="FFFF00")
        ws.column_dimensions[chr(65+col)].width = colsizes[col]
    r = 2
    for m in mlist:
        ws.cell(row=r,column=1).value = v(m,'first_name')+" "+v(m,'last_name')
        ws.cell(row=r,column=2).value = v(m,'unit')
        ws.cell(row=r,column=3).value = v(m,'contact_home_phone')
        ws.cell(row=r,column=4).value = v(m,'contact_mobile_phone')
        ws.cell(row=r,column=5).value = v(m,'contact_mobile_mum')
        ws.cell(row=r,column=6).value = v(m,'contact_mobile_dad')
        r += 1
    save_file("Telefonlista.xlsx",save_virtual_workbook(wb))

# Droxbox upload function

# import dropbox
# 
# DBX_OAUTHKEY = os.getenv('DBX_OAUTHKEY', 'NO DEFAULT!')
# DBX_BASEDIR = "/Aktuella kontakt- och e-postlistor/"
# GDRIVE_BASEDIR = "1HGb1JHNFx5-GxrTbzyK7RpQr5DID8FXS"


AVDELNINGSDOKUMENT_ID = "0B4TmjLu89np8NkRoLWd3VVU2ZUE"

from .gdrive import GDrive
import json

def save_file(fname, data):
    # NO MORE DROPBOX!!!
#     dbx = dropbox.Dropbox(DBX_OAUTHKEY)
#     dbx.files_upload(data, DBX_BASEDIR+fname, dropbox.files.WriteMode.overwrite, mute=True)
    gservice = GDrive(json.loads(os.getenv("service_account_info", "")))
    epostlist_dir = gservice.find_file(AVDELNINGSDOKUMENT_ID,"Aktuella kontakt- och e-postlistor")[0]['id']
    gservice.write_file(fname,epostlist_dir,data)
# BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# def save_file(fname,data):
#     f = open(BASE_DIR+"/../TEMP/"+fname,"wb")
#     f.write(data)
#     f.close()

# Scoutnet download function

import requests

def get_memdata():
    dataurl = "https://{authkey}@www.scoutnet.se/api/group/memberlist"
    auth = os.getenv('scoutnet_apikey','NO DEFAULT!')
    s = requests.Session()
    r = s.get(dataurl.format(authkey=auth))
    return r.json()

# dataurl = "https://www.scoutnet.se/reports/groups/members/group_id/784/download/true/format/json"
# loginurl = "https://www.scoutnet.se/login"
# auth = {'signin[username]': os.getenv('SCOUTNET_UID','hakan@violaberg.nu'), 'signin[password]': os.getenv('SCOUTNET_PWD','NO DEFAULT!')}
#   
# def get_memdata():
#     s = requests.Session()
#     r = s.get(dataurl)
#     if r.status_code != 200:
#         r = s.post(loginurl,data=auth)  # Need to login
#         if r.status_code != 200:
#             raise Exception('Bad Scoutnet credentials')
#     return r.json()
 
# BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# import json
# def get_memdata():
#     return json.load(open(BASE_DIR+"/alla.json","r"))
# }

